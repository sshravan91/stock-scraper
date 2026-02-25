import os
import re
from typing import Any

import xlrd
import xlwt
from openpyxl import load_workbook


OUTPUT_COLUMNS = [
    "Scheme name",
    "Category",
    "NAV",
    "AUM",
    "TER",
    "1 Yr Rtn",
    "3 Yr Rtn",
    "5 Yr Rtn",
    "10 Yr Rtn",
    "Volatility",
    "Standard Deviation",
    "Sharpe Ratio",
    "Beta",
    "Alpha",
    "Mean",
    "Sortino Ratio",
    "Up Market Capture Ratio",
    "Down Market Capture Ratio",
    "Maximum Drawdown",
    "R-Squared",
    "Information Ratio",
    "Treynor Ratio",
]


def _normalize(text: Any) -> str:
    s = str(text or "").strip().lower()
    s = s.replace("&", " and ")
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s


def _load_rows(path: str) -> tuple[list[str], list[list[Any]]]:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        wb = xlrd.open_workbook(path)
        sh = wb.sheet_by_index(0)
        rows = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    else:
        wb = load_workbook(filename=path, data_only=True, read_only=True)
        ws = wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]

    header_row_idx = 0
    for i, row in enumerate(rows[:60]):
        tokens = " ".join(str(v or "").lower() for v in row)
        if "scheme" in tokens and "category" in tokens:
            header_row_idx = i
            break

    headers = [str(v or "").strip() for v in rows[header_row_idx]]
    data_rows = rows[header_row_idx + 1 :]
    return headers, data_rows


def _build_col_index(headers: list[str]) -> dict[str, int]:
    nheaders = [_normalize(h) for h in headers]

    def pick(*aliases: str) -> int:
        norm_aliases = [_normalize(a) for a in aliases]
        for alias in norm_aliases:
            for i, h in enumerate(nheaders):
                if h == alias:
                    return i
        for alias in norm_aliases:
            for i, h in enumerate(nheaders):
                if alias in h or h in alias:
                    return i
        return -1

    return {
        "scheme": pick("Scheme Name", "Scheme", "Fund Name", "Fund"),
        "category": pick("Category"),
        "nav": pick("NAV", "Nav"),
        "aum": pick("AUM", "AUM (Cr)", "AUM (in Cr)"),
        "ter": pick("TER", "Expense Ratio"),
        "r1": pick("1 Year Return", "1 Yr Return", "1 Year Rtn", "1 Yr Rtn"),
        "r3": pick("3 Year Return", "3 Yr Return", "3 Year Rtn", "3 Yr Rtn"),
        "r5": pick("5 Year Return", "5 Yr Return", "5 Year Rtn", "5 Yr Rtn"),
        "r10": pick("10 Year Return", "10 Yr Return", "10 Year Rtn", "10 Yr Rtn"),
        "volatility": pick("Volatility"),
        "stddev": pick("Standard Deviation"),
        "sharpe": pick("Sharpe Ratio", "Sharp Ratio"),
        "beta": pick("Beta"),
        "alpha": pick("Alpha"),
        "mean": pick("Mean"),
        "sortino": pick("Sortino Ratio"),
        "up_capture": pick("Up Market Capture Ratio", "Up Market Capture\nRatio"),
        "down_capture": pick("Down Market Capture Ratio", "Down Market Capture\nRatio"),
        "max_dd": pick("Maximum Drawdown"),
        "r_squared": pick("R-Squared", "R Squared"),
        "info_ratio": pick("Information Ratio"),
        "treynor": pick("Treynor Ratio"),
    }


def _value(row: list[Any], idx: int) -> Any:
    if idx < 0 or idx >= len(row):
        return ""
    v = row[idx]
    if v is None:
        return ""
    return v


def _scheme_key(name: Any) -> str:
    return _normalize(name)


def consolidate_mft_returns(
    trailing_path: str,
    risk_path: str,
    output_path: str | None = None,
) -> str:
    if not output_path:
        output_path = os.path.join(os.getcwd(), "consolidated-mft-returns.xls")

    trailing_headers, trailing_rows = _load_rows(trailing_path)
    risk_headers, risk_rows = _load_rows(risk_path)

    t_idx = _build_col_index(trailing_headers)
    r_idx = _build_col_index(risk_headers)

    trailing_by_scheme: dict[str, dict[str, Any]] = {}
    for row in trailing_rows:
        scheme = _value(row, t_idx["scheme"])
        key = _scheme_key(scheme)
        if not key:
            continue
        trailing_by_scheme[key] = {
            "Scheme name": scheme,
            "Category": _value(row, t_idx["category"]),
            "NAV": _value(row, t_idx["nav"]),
            "AUM": _value(row, t_idx["aum"]),
            "TER": _value(row, t_idx["ter"]),
            "1 Yr Rtn": _value(row, t_idx["r1"]),
            "3 Yr Rtn": _value(row, t_idx["r3"]),
            "5 Yr Rtn": _value(row, t_idx["r5"]),
            "10 Yr Rtn": _value(row, t_idx["r10"]),
        }

    risk_by_scheme: dict[str, dict[str, Any]] = {}
    for row in risk_rows:
        scheme = _value(row, r_idx["scheme"])
        key = _scheme_key(scheme)
        if not key:
            continue
        risk_by_scheme[key] = {
            "Volatility": _value(row, r_idx["volatility"]),
            "Standard Deviation": _value(row, r_idx["stddev"]),
            "Sharpe Ratio": _value(row, r_idx["sharpe"]),
            "Beta": _value(row, r_idx["beta"]),
            "Alpha": _value(row, r_idx["alpha"]),
            "Mean": _value(row, r_idx["mean"]),
            "Sortino Ratio": _value(row, r_idx["sortino"]),
            "Up Market Capture Ratio": _value(row, r_idx["up_capture"]),
            "Down Market Capture Ratio": _value(row, r_idx["down_capture"]),
            "Maximum Drawdown": _value(row, r_idx["max_dd"]),
            "R-Squared": _value(row, r_idx["r_squared"]),
            "Information Ratio": _value(row, r_idx["info_ratio"]),
            "Treynor Ratio": _value(row, r_idx["treynor"]),
        }

    # Keep trailing report as primary universe, merge risk metrics where available.
    merged_rows: list[dict[str, Any]] = []
    for key, trow in trailing_by_scheme.items():
        row = {col: "" for col in OUTPUT_COLUMNS}
        row.update(trow)
        row.update(risk_by_scheme.get(key, {}))
        merged_rows.append(row)

    wb = xlwt.Workbook()
    ws = wb.add_sheet("Consolidated Returns")

    for c, col in enumerate(OUTPUT_COLUMNS):
        ws.write(0, c, col)

    for r, row in enumerate(merged_rows, start=1):
        for c, col in enumerate(OUTPUT_COLUMNS):
            ws.write(r, c, row.get(col, ""))

    wb.save(output_path)
    return os.path.abspath(output_path)

