import concurrent.futures
import csv
import requests
import yaml
import json
from bs4 import BeautifulSoup
from datetime import datetime
from collections import defaultdict
import threading
import re
import argparse
import os
from openpyxl import load_workbook
import xlrd  # for legacy .xls files

funds_with_no_data = []

# Mappings for enrichment
MFT_AK_TO_KEY = {}
MFT_AK_TO_AMFI = {}
MFT_AK_LIST = []
MFT_CATEGORIES = []
RISK_METRICS_BY_MFTOOLS_KEY = {}

def load_mftools_mapping(json_path):
    try:
        global MFT_AK_LIST, MFT_CATEGORIES
        # reset caches
        MFT_AK_TO_KEY.clear()
        MFT_AK_TO_AMFI.clear()
        MFT_AK_LIST = []
        with open(json_path, "r") as f:
            payload = json.load(f)
        for fnd in payload.get("funds", []):
            ak = fnd.get("akKey")
            mk = fnd.get("mftools_key")
            amfi = fnd.get("amfiKey")
            if isinstance(ak, str) and isinstance(mk, str) and mk.strip():
                MFT_AK_TO_KEY[ak] = mk.strip()
            if isinstance(ak, str) and isinstance(amfi, str) and amfi.strip():
                MFT_AK_TO_AMFI[ak] = amfi.strip()
            if isinstance(ak, str) and ak.strip():
                MFT_AK_LIST.append(ak.strip())
        MFT_CATEGORIES = payload.get("categories") or []
    except Exception as ex:
        print(f"Failed to load mftools mapping from {json_path}: {ex}")

def load_risk_ratios(path):
    try:
        ext = os.path.splitext(path)[1].lower()
        required = ["Volatility", "Sharpe Ratio", "Beta", "Alpha", "Mean", "Sortino Ratio", "Up Market Capture\nRatio", "Down Market Capture\nRatio", "Maximum Drawdown", "R-Squared", "Information Ratio"]
        if ext == ".xls":
            if xlrd is None:
                raise RuntimeError("xlrd is required to read .xls files. Install with 'pip install xlrd'.")
            wb = xlrd.open_workbook(path)
            sh = wb.sheet_by_index(0)
            # autodetect header row by scanning first 200 rows for a row that contains expected header cues
            header_row_idx = None
            headers = []
            for rr in range(min(sh.nrows, 200)):
                row_vals = []
                for cc in range(sh.ncols):
                    val = sh.cell_value(rr, cc)
                    row_vals.append(str(val).strip() if val is not None else "")
                joined = " ".join(v.lower() for v in row_vals)
                if "scheme name" in joined and "category" in joined:
                    header_row_idx = rr
                    headers = row_vals
                    break
            if header_row_idx is None:
                header_row_idx = 0
                headers = []
                for c in range(sh.ncols):
                    headers.append(str(sh.cell_value(0, c)).strip() if sh.cell_value(0, c) is not None else "")
            # build fuzzy header index mapping
            lower_headers = [h.lower() for h in headers]
            h_idx = {}
            for k in required:
                lk = k.lower()
                idx = None
                for i, hv in enumerate(lower_headers):
                    if lk in hv or hv in lk:
                        idx = i
                        break
                if idx is not None:
                    h_idx[k] = idx
            # iterate rows after detected header
            for r in range(header_row_idx + 1, sh.nrows):
                key = sh.cell_value(r, 0)
                if key in (None, ""):
                    continue
                key = str(key).strip()
                metrics = {}
                for k, idx in h_idx.items():
                    if idx is not None and idx < sh.ncols:
                        metrics[k] = sh.cell_value(r, idx)
                if key:
                    RISK_METRICS_BY_MFTOOLS_KEY[key] = metrics
        else:
            wb = load_workbook(filename=path, data_only=True, read_only=True)
            ws = wb.active
            # buffer first 200 rows to locate header row
            buffered = []
            it = ws.iter_rows(values_only=True)
            for _ in range(200):
                try:
                    buffered.append(next(it))
                except StopIteration:
                    break
            header_row = None
            header_row_idx = None
            for i, row in enumerate(buffered):
                if not row:
                    continue
                joined = " ".join((str(v).strip().lower() if v is not None else "") for v in row)
                if "volatility" in joined and "sharpe" in joined:
                    header_row = row
                    header_row_idx = i
                    break
            if header_row is None:
                if buffered:
                    header_row = buffered[0]
                    header_row_idx = 0
                else:
                    header_row = []
                    header_row_idx = 0
            headers = [str(h).strip() if h is not None else "" for h in header_row]
            lower_headers = [h.lower() for h in headers]
            h_idx = {}
            for k in required:
                lk = k.lower()
                idx = None
                for i, hv in enumerate(lower_headers):
                    if lk in hv or hv in lk:
                        idx = i
                        break
                if idx is not None:
                    h_idx[k] = idx
            # process buffered rows after header
            for r in buffered[header_row_idx + 1:]:
                if not r:
                    continue
                key = r[0]
                if key is None or str(key).strip() == "":
                    continue
                key = str(key).strip()
                metrics = {}
                for k, idx in h_idx.items():
                    if idx is not None and idx < len(r):
                        metrics[k] = r[idx]
                if key:
                    RISK_METRICS_BY_MFTOOLS_KEY[key] = metrics
            # continue with the rest of rows from iterator
            for r in it:
                if not r:
                    continue
                key = r[0]
                if key is None or str(key).strip() == "":
                    continue
                key = str(key).strip()
                metrics = {}
                for k, idx in h_idx.items():
                    if idx is not None and idx < len(r):
                        metrics[k] = r[idx]
                if key:
                    RISK_METRICS_BY_MFTOOLS_KEY[key] = metrics
    except Exception as ex:
        print(f"Failed to load risk ratios from {path}: {ex}")

def enrich_from_mftools(valueDict, ak_name):
    try:
        mkey = MFT_AK_TO_KEY.get(ak_name)
        if isinstance(mkey, str) and mkey:
            metrics = RISK_METRICS_BY_MFTOOLS_KEY.get(mkey)
            if metrics:
                for k, v in metrics.items():
                    if v is not None:
                        valueDict[k] = v
    except Exception as ex:
        print(f"Risk metrics enrichment failed for {ak_name}: {ex}")

def get_stock_prices(tickers):
    trendDetails = list()
    futures = list()
    symbols_no_data = []

    # Number of processes in the process pool
    num_processes = 100  # You can adjust this based on your system resources

    # Use ThreadPoolExecutor for I/O-bound web requests
    with concurrent.futures.ThreadPoolExecutor(max_workers=min(16, len(tickers))) as executor:
        for count, ticker in enumerate(tickers):
            print(count + 1, ticker)
            futures.append(executor.submit(get_stock_info, ticker))

        for future in concurrent.futures.as_completed(futures):
            has_data, value = future.result()
            if has_data:
                trendDetails.append(value)
            else:
                symbols_no_data.append(value)

    # Update the global funds_with_no_data
    global funds_with_no_data
    funds_with_no_data = symbols_no_data

    return trendDetails

def get_stock_info(symbol):
    ak_base_url = "https://www.advisorkhoj.com/mutual-funds-research/"
    # symbol can be of the form "DisplayName:slug". Extract both parts.
    sym0 = symbol
    sym1 = None
    if ':' in symbol:
        sym0, sym1 = symbol.split(':', 1)
    url = f"{ak_base_url}{sym0}"

    ak_response = requests.get(url, headers={}, timeout=20)

    if ak_response.status_code == 200:
        valueDict = {}

        cagr_mapping = {'scheme_inception_returns': 'CAGR Since Inception',
                        'scheme_1yr_returns': '1 Year CAGR',
                        'scheme_3yr_returns': '3 Years CAGR',
                        'scheme_5yr_returns': '5 Years CAGR',
                        'scheme_10yr_returns': '10 Years CAGR',
                        'category_1yr_returns': '1 Year Category CAGR',
                        'category_3yr_returns': '3 Years Category CAGR',
                        'category_5yr_returns': '5 Years Category CAGR',
                        'category_10yr_returns': '10 Years Category CAGR',
                        'benchmark_1yr_returns': '1 Year Benchmark CAGR',
                        'benchmark_3yr_returns': '3 Years Benchmark CAGR',
                        'benchmark_5yr_returns': '5 Years Benchmark CAGR',
                        'benchmark_10yr_returns': '10 Years Benchmark CAGR',
                        'scheme_nav': 'NAV',
                        'scheme_benchmark': 'Benchmark Type'}

        ak_response_txt = ak_response.text
        for key, index in cagr_mapping.items():
            value = extract_using_regex(ak_response_txt, key)
            if value:
                valueDict[index] = value

        soup = BeautifulSoup(ak_response_txt, 'html.parser')

        # Fallback: extract NAV from DOM if not found via JS vars
        if 'NAV' not in valueDict:
            nav_label = soup.find('div', class_='nav-cagr-label')
            if nav_label and 'NAV as on' in nav_label.get_text():
                nav_value = nav_label.find_next('h4')
                if nav_value:
                    valueDict['NAV'] = nav_value.get_text(strip=True).replace('₹', '').replace(',', '')

        context_mapping = {'Category: ': 'Category',
                           'TER:': 'TER',
                           'Total Assets:': 'Total Assets (in Cr)',
                           'Launch Date:': 'Launch Date',
                           'Turn over:': 'Turn over (%)',
                           'Standard Deviation': 'Standard Deviation',
                           'Alpha': 'Alpha',
                           'Beta': 'Beta',
                           'Sharpe Ratio': 'Sharpe Ratio'}

        # retrieve market cap distributions
        # mkt_cap_dist_types = ['Small Cap', 'Others', 'Large Cap', 'Mid Cap']
        # for mkt_cap_div in soup.find_all('div', class_='flex-div'):
        #     label_p = mkt_cap_div.find('p', class_='font12 text-left')
        #     if not label_p:
        #         continue
        #     category = label_p.get_text(strip=True)
        #     if category in mkt_cap_dist_types:
        #         # the bar container has a style like "width:12.85%" and inner div with title="12.85%"
        #         bar_container = mkt_cap_div.find('div', style=re.compile(r'width:\s*\d'))
        #         if bar_container and bar_container.div and bar_container.div.has_attr('title'):
        #             valueDict[category] = bar_container.div['title'].strip()

        sch_over_table_keys = {'Category: ',
                               'TER:',
                               'Turn over:',
                               'Total Assets:',
                               'Launch Date:'}
        tables = soup.find_all('table', class_='sch_over_table')
        for index, table in enumerate(tables, start=1):
            rows = table.find_all('tr')
            for row in rows:
                subrow = row.find('td').text.strip()
                for key in sch_over_table_keys:
                    if key in subrow:
                        if key == 'TER:':
                            valueDict[context_mapping[key]] = subrow.replace(key, '').strip().split(" As on ")[0]
                        elif key == 'Total Assets:':
                            valueDict[context_mapping[key]] = subrow.replace(key, '').strip().split(" Cr As on ")[0]
                        elif key == 'Turn over:':
                            result = subrow.replace(key, '').strip().split("|")[0]
                            valueDict[context_mapping[key]] = result.strip() if result else result
                        else:
                            valueDict[context_mapping[key]] = subrow.replace(key, '').strip()

        adv_table_keys = {'Standard Deviation', 'Sharpe Ratio', 'Alpha', 'Beta'}
        tables = soup.find_all('table', class_='adv-table table table-striped')
        for index, table in enumerate(tables, start=1):
            rows = table.find_all('tr')
            for row in rows:
                subrow = row.find('td')
                if subrow:
                    cleanedsubrow = subrow.text.strip()
                    for key in adv_table_keys:
                        if key in cleanedsubrow:
                            valueDict[context_mapping[key]] = row.find_next('td', {'class': 'text-center'}).text.strip()

        print("Finished parsing " + url)

        if bool(valueDict):
            valueDict['Fund'] = sym0
            # Enrich with mftools risk ratios if available before potential early return
            enrich_from_mftools(valueDict, sym0)
            # if sym1 is None:
            #     return (True, valueDict)
        else:
            print(f"\033[91m{sym0} has no data\033[0m")
            return (False, sym0)
    else:
        print(f"\033Failed for {url}\033[0m")
        return (False, sym0)

    # Resolve scheme code via amfiKey mapping from mftools JSON (no web search)
    scheme_code = MFT_AK_TO_AMFI.get(sym0)
    if isinstance(scheme_code, str) and scheme_code.strip():
        valueDict['Scheme Code'] = scheme_code.strip()

    scheme_code = valueDict.get('Scheme Code')
    if isinstance(scheme_code, str) and scheme_code.strip():
        try:
            groww_page_url = f"https://groww.in/v1/api/data/mf/web/v1/scheme/portfolio/{scheme_code}/stats"
            gp_resp = requests.get(groww_page_url, timeout=20)
            if gp_resp.status_code == 200:
                data = json.loads(gp_resp.text)
                valueDict['P/E Ratio'] = data.get("pe")
                valueDict['P/B Ratio'] = data.get("pb")
        except Exception:
            # Do not fail overall parsing if Groww page structure changes
            pass

    # Final enrichment with risk metrics (if not already applied)
    enrich_from_mftools(valueDict, sym0)
    return (True, valueDict)

def extract_using_regex(input_string, key):
    # Match JS assignments like: var key = 'value'; or key="value"
    pattern = rf"\b{re.escape(key)}\s*=\s*['\"]([^'\"]+)['\"]"
    match = re.search(pattern, input_string)
    if match:
        return match.group(1).strip()
    return None

def export_to_file(data):
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    csv_file_path = f"fund-stats_{timestamp}.csv"

    column_order = ['Fund',
                   'Category',
                   'Scheme Code',
                   'Launch Date',
                   'Total Assets (in Cr)',
                   'TER',
                   'Turn over (%)',
                   'CAGR Since Inception',
                   '1 Year CAGR',
                   '1 Year Category CAGR',
                   '1 Year Benchmark CAGR',
                   '3 Years CAGR',
                   '3 Years Category CAGR',
                   '3 Years Benchmark CAGR',
                   '5 Years CAGR',
                   '5 Years Category CAGR',
                   '5 Years Benchmark CAGR',
                   '10 Years CAGR',
                   '10 Years Category CAGR',
                   '10 Years Benchmark CAGR',
                   'Benchmark Type',
                   'NAV',
                   'Alpha',
                   'Beta',
                   'Standard Deviation',
                   'Sharpe Ratio',
                   "Volatility",
                   "Mean",
                   "Sortino Ratio", 
                   "Up Market Capture\nRatio", 
                   "Down Market Capture\nRatio", 
                   "Maximum Drawdown", 
                   "R-Squared",
                   "Information Ratio",
                   'P/E Ratio',
                   'P/B Ratio'
                   ]

    category_order = MFT_CATEGORIES or []

    fundsByType = {}
    for fund_data in data:
        if 'Category' in fund_data:
            category = fund_data.get('Category')
            fundStats = []
            if category in fundsByType:
                fundStats = fundsByType.get(category)
            fundStats.append([fund_data.get(column, '') for column in column_order])
            fundsByType[category] = fundStats
        else:
            print("No category for fund " + str(fund_data) + " hence skipping.")

    with open(csv_file_path, 'w', newline='') as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(column_order)

        for category in category_order:
            if category in fundsByType.keys():
                writer.writerow([category])
                writer.writerows(fundsByType[category])

def extract_data_from_yaml(property):
    with open('fundslist.yaml', 'r') as file:
        data = yaml.safe_load(file)

    return data[property]

def build_fund_records_from_yaml():
    """
    Build seed fund records from fundslist.yaml where each entry may be in the form:
      "AK-Display-Name" or "AK-Display-Name:groww-slug"
    Returns a list of dicts: [{akKey, growwKey, amfiKey}, ...]
    """
    funds_list = extract_data_from_yaml('funds')
    records = []
    for item in funds_list:
        ak_key = item
        groww_key = None
        if ':' in item:
            ak_key, groww_key = item.split(':', 1)
        records.append({'akKey': ak_key, 'growwKey': groww_key, 'amfiKey': None})
    return records

def enrich_fund_records_with_amfi(records, parsed_fund_data):
    """
    For each parsed fund data item, if it has a Scheme Code, attach it as amfiKey
    to the corresponding record matched by akKey (valueDict['Fund']).
    """
    try:
        # index = {rec['akKey']: rec for rec in records}
        index = {}
        for rec in records:
            index[rec['akKey']] = rec

        for fd in parsed_fund_data:
            ak = fd.get('Fund')
            amfi = fd.get('Scheme Code')
            category = fd.get('Category')
            if ak and amfi and ak in index:
                index[ak]['amfiKey'] = amfi
                index[ak]['category'] = category
        return index
    except Exception as ex:
        print(ex)

def export_funds_categories_json(records, categories, output_path='funds_and_categories.json'):
    """
    Export the required JSON with top-level keys 'funds' and 'categories'.
    """
    payload = {
        'funds': records,
        'categories': categories
    }
    with open(output_path, 'w') as f:
        json.dump(payload, f, indent=2)

if __name__ == "__main__":
    # parse arguments
    parser = argparse.ArgumentParser(description="Advisor parser with risk ratios enrichment")
    script_dir = os.path.dirname(os.path.abspath(__file__))
    default_mftools_json = os.path.join(script_dir, "funds_and_categories_with_mftools.json")
    default_riskratios = os.path.join(script_dir, "risk-ratios.xls")
    parser.add_argument("--risk-ratios", default=default_riskratios, help="Path to risk ratios Excel (.xls or .xlsx)")
    parser.add_argument("--mftools-json", default=default_mftools_json, help="Path to funds_and_categories_with_mftools.json (defaults to file alongside this script)")
    args = parser.parse_args()

    # load mappings
    load_mftools_mapping(args.mftools_json)
    load_risk_ratios(args.risk_ratios)

    # extract fund names from JSON (akKey list)
    funds = MFT_AK_LIST or []

    # extract data for funds
    extracted_data = get_stock_prices(funds)

    # data_sorted_by_alpha = sorted(extracted_data, key=lambda x: (print(x) or float(x['Alpha'])) if x['Alpha'] and x['Alpha'] != '-' else float('-inf'), reverse=True)
    print(f"\033[91m{len(funds_with_no_data)} funds have no data. These are, {funds_with_no_data}.\033[0m")

    # export CSV file (existing behavior)
    export_to_file(extracted_data)
