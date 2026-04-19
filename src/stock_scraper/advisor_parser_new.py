import concurrent.futures
import csv
import requests
import json
from datetime import datetime
import re
import argparse
import os
from openpyxl import load_workbook
import xlrd  # for legacy .xls files

funds_with_no_data = []

# Mappings for enrichment
MFT_AK_TO_KEY = {}
MFT_AK_TO_AMFI = {}
MFT_AK_TO_BENCHMARK = {}
MFT_AK_TO_CATEGORY = {}
MFT_AK_LIST = []
MFT_CATEGORIES = []
RISK_METRICS_BY_MFTOOLS_KEY = {}
CATEGORY_RETURNS_BY_NAME = {}
BENCHMARK_RETURNS_BY_NAME = {}


def normalize_key(value):
    text = str(value or "").strip().lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text

def load_mftools_mapping(json_path):
    try:
        global MFT_AK_LIST, MFT_CATEGORIES
        # reset caches
        MFT_AK_TO_KEY.clear()
        MFT_AK_TO_AMFI.clear()
        MFT_AK_TO_BENCHMARK.clear()
        MFT_AK_TO_CATEGORY.clear()
        MFT_AK_LIST = []
        with open(json_path, "r") as f:
            payload = json.load(f)
        for fnd in payload.get("funds", []):
            ak = fnd.get("akKey")
            mk = fnd.get("mftools_key")
            amfi = fnd.get("amfiKey")
            benchmark = fnd.get("benchmark")
            category = fnd.get("category")
            if isinstance(ak, str) and isinstance(mk, str) and mk.strip():
                MFT_AK_TO_KEY[ak] = mk.strip()
            if isinstance(ak, str) and isinstance(amfi, str) and amfi.strip():
                MFT_AK_TO_AMFI[ak] = amfi.strip()
            if isinstance(ak, str) and isinstance(benchmark, str) and benchmark.strip():
                MFT_AK_TO_BENCHMARK[ak] = benchmark.strip()
            if isinstance(ak, str) and isinstance(category, str) and category.strip():
                MFT_AK_TO_CATEGORY[ak] = category.strip()
            if isinstance(ak, str) and ak.strip():
                MFT_AK_LIST.append(ak.strip())
        MFT_CATEGORIES = payload.get("categories") or []
    except Exception as ex:
        print(f"Failed to load mftools mapping from {json_path}: {ex}")


def _load_rows_generic(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        wb = xlrd.open_workbook(path)
        sh = wb.sheet_by_index(0)
        return [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    wb = load_workbook(filename=path, data_only=True, read_only=True)
    ws = wb.active
    return [list(r) for r in ws.iter_rows(values_only=True)]


def load_category_returns(path):
    CATEGORY_RETURNS_BY_NAME.clear()
    try:
        rows = _load_rows_generic(path)
        if not rows:
            return
        header_idx = 0
        for i, row in enumerate(rows[:80]):
            tokens = " ".join(str(v or "").lower() for v in row)
            if "category name" in tokens and "1 yr" in tokens:
                header_idx = i
                break
        headers = [str(v or "").strip() for v in rows[header_idx]]
        nheaders = [normalize_key(h) for h in headers]

        def pick(*aliases):
            ns = [normalize_key(a) for a in aliases]
            for a in ns:
                for i, h in enumerate(nheaders):
                    if h == a:
                        return i
            for a in ns:
                for i, h in enumerate(nheaders):
                    if a in h or h in a:
                        return i
            return -1

        idx_name = pick("Category Name", "Category")
        idx_1 = pick("1 Yr", "1 Year", "1 Yr Return")
        idx_3 = pick("3 Yrs", "3 Yr", "3 Year")
        idx_5 = pick("5 Yrs", "5 Yr", "5 Year")
        idx_10 = pick("10 Yrs", "10 Yr", "10 Year")
        for row in rows[header_idx + 1:]:
            if idx_name < 0 or idx_name >= len(row):
                continue
            name = str(row[idx_name] or "").strip()
            if not name:
                continue
            CATEGORY_RETURNS_BY_NAME[normalize_key(name)] = {
                "1 Year Category CAGR": row[idx_1] if 0 <= idx_1 < len(row) and row[idx_1] is not None else "",
                "3 Years Category CAGR": row[idx_3] if 0 <= idx_3 < len(row) and row[idx_3] is not None else "",
                "5 Years Category CAGR": row[idx_5] if 0 <= idx_5 < len(row) and row[idx_5] is not None else "",
                "10 Years Category CAGR": row[idx_10] if 0 <= idx_10 < len(row) and row[idx_10] is not None else "",
            }
    except Exception as ex:
        print(f"Failed to load category monitor from {path}: {ex}")


def load_benchmark_returns(path):
    BENCHMARK_RETURNS_BY_NAME.clear()
    try:
        rows = _load_rows_generic(path)
        if not rows:
            return
        header_idx = 0
        for i, row in enumerate(rows[:80]):
            tokens = " ".join(str(v or "").lower() for v in row)
            if "benchmark name" in tokens and "1 yr" in tokens:
                header_idx = i
                break
        headers = [str(v or "").strip() for v in rows[header_idx]]
        nheaders = [normalize_key(h) for h in headers]

        def pick(*aliases):
            ns = [normalize_key(a) for a in aliases]
            for a in ns:
                for i, h in enumerate(nheaders):
                    if h == a:
                        return i
            for a in ns:
                for i, h in enumerate(nheaders):
                    if a in h or h in a:
                        return i
            return -1

        idx_name = pick("Benchmark Name", "Benchmark")
        idx_1 = pick("1 Yr", "1 Year", "1 Yr Return")
        idx_3 = pick("3 Yrs", "3 Yr", "3 Year")
        idx_5 = pick("5 Yrs", "5 Yr", "5 Year")
        idx_10 = pick("10 Yrs", "10 Yr", "10 Year")
        for row in rows[header_idx + 1:]:
            if idx_name < 0 or idx_name >= len(row):
                continue
            name = str(row[idx_name] or "").strip()
            if not name:
                continue
            BENCHMARK_RETURNS_BY_NAME[normalize_key(name)] = {
                "1 Year Benchmark CAGR": row[idx_1] if 0 <= idx_1 < len(row) and row[idx_1] is not None else "",
                "3 Years Benchmark CAGR": row[idx_3] if 0 <= idx_3 < len(row) and row[idx_3] is not None else "",
                "5 Years Benchmark CAGR": row[idx_5] if 0 <= idx_5 < len(row) and row[idx_5] is not None else "",
                "10 Years Benchmark CAGR": row[idx_10] if 0 <= idx_10 < len(row) and row[idx_10] is not None else "",
            }
    except Exception as ex:
        print(f"Failed to load benchmark returns from {path}: {ex}")


def enrich_category_and_benchmark_returns(valueDict, ak_name):
    mapped_category = MFT_AK_TO_CATEGORY.get(ak_name)
    if mapped_category:
        valueDict["Category"] = mapped_category
        category_returns = CATEGORY_RETURNS_BY_NAME.get(normalize_key(mapped_category))
        if category_returns:
            valueDict.update(category_returns)

    mapped_benchmark = MFT_AK_TO_BENCHMARK.get(ak_name)
    if mapped_benchmark:
        valueDict["Benchmark Type"] = mapped_benchmark
        benchmark_returns = BENCHMARK_RETURNS_BY_NAME.get(normalize_key(mapped_benchmark))
        if benchmark_returns:
            valueDict.update(benchmark_returns)

def load_risk_ratios(path):
    try:
        ext = os.path.splitext(path)[1].lower()
        required = [
            "Category",
            "NAV",
            "AUM",
            "TER",
            "1 Yr Rtn",
            "3 Yr Rtn",
            "5 Yr Rtn",
            "10 Yr Rtn",
            "Volatility",
            "Standard Deviation",
            "Sharpe Ratio",
            "Beta",
            "Alpha",
            "Mean",
            "Sortino Ratio",
            "Up Market Capture Ratio",
            "Down Market Capture Ratio",
            "Maximum Drawdown",
            "R-Squared",
            "Information Ratio",
        ]

        def pick(headers, *aliases):
            normalized_headers = [normalize_key(h) for h in headers]
            normalized_aliases = [normalize_key(a) for a in aliases]
            for alias in normalized_aliases:
                for i, h in enumerate(normalized_headers):
                    if h == alias:
                        return i
            for alias in normalized_aliases:
                for i, h in enumerate(normalized_headers):
                    if alias in h or h in alias:
                        return i
            return -1

        def build_stats_row(row, idx):
            def val(k):
                i = idx.get(k, -1)
                if i < 0 or i >= len(row):
                    return ""
                v = row[i]
                return "" if v is None else v

            return {
                "Category": val("Category"),
                "NAV": val("NAV"),
                "Total Assets (in Cr)": val("AUM"),
                "TER": val("TER"),
                "1 Year CAGR": val("1 Yr Rtn"),
                "3 Years CAGR": val("3 Yr Rtn"),
                "5 Years CAGR": val("5 Yr Rtn"),
                "10 Years CAGR": val("10 Yr Rtn"),
                "Volatility": val("Volatility"),
                "Standard Deviation": val("Standard Deviation"),
                "Sharpe Ratio": val("Sharpe Ratio"),
                "Beta": val("Beta"),
                "Alpha": val("Alpha"),
                "Mean": val("Mean"),
                "Sortino Ratio": val("Sortino Ratio"),
                "Up Market Capture\nRatio": val("Up Market Capture Ratio"),
                "Down Market Capture\nRatio": val("Down Market Capture Ratio"),
                "Maximum Drawdown": val("Maximum Drawdown"),
                "R-Squared": val("R-Squared"),
                "Information Ratio": val("Information Ratio"),
            }

        if ext == ".xls":
            if xlrd is None:
                raise RuntimeError("xlrd is required to read .xls files. Install with 'pip install xlrd'.")
            wb = xlrd.open_workbook(path)
            sh = wb.sheet_by_index(0)
            # autodetect header row by scanning first 200 rows for a row that contains expected header cues
            header_row_idx = None
            headers = []
            for rr in range(min(sh.nrows, 200)):
                row_vals = []
                for cc in range(sh.ncols):
                    val = sh.cell_value(rr, cc)
                    row_vals.append(str(val).strip() if val is not None else "")
                joined = " ".join(v.lower() for v in row_vals)
                if "scheme name" in joined and "category" in joined:
                    header_row_idx = rr
                    headers = row_vals
                    break
            if header_row_idx is None:
                header_row_idx = 0
                headers = []
                for c in range(sh.ncols):
                    headers.append(str(sh.cell_value(0, c)).strip() if sh.cell_value(0, c) is not None else "")
            h_idx = {
                "Category": pick(headers, "Category"),
                "NAV": pick(headers, "NAV", "Nav"),
                "AUM": pick(headers, "AUM", "AUM (Crore)", "AUM (Cr)", "AUM (in Cr)"),
                "TER": pick(headers, "TER", "TER (%)"),
                "1 Yr Rtn": pick(headers, "1 Yr Rtn", "1 Year Rtn", "1 Yr Rtn (%)", "1 Year Rtn (%)"),
                "3 Yr Rtn": pick(headers, "3 Yr Rtn", "3 Year Rtn", "3 Yrs Rtn (%)", "3 Yr Rtn (%)"),
                "5 Yr Rtn": pick(headers, "5 Yr Rtn", "5 Year Rtn", "5 Yrs Rtn (%)", "5 Yr Rtn (%)"),
                "10 Yr Rtn": pick(headers, "10 Yr Rtn", "10 Year Rtn", "10 Yrs Rtn (%)", "10 Yr Rtn (%)"),
                "Volatility": pick(headers, "Volatility"),
                "Standard Deviation": pick(headers, "Standard Deviation"),
                "Sharpe Ratio": pick(headers, "Sharpe Ratio", "Sharp Ratio"),
                "Beta": pick(headers, "Beta"),
                "Alpha": pick(headers, "Alpha"),
                "Mean": pick(headers, "Mean"),
                "Sortino Ratio": pick(headers, "Sortino Ratio"),
                "Up Market Capture Ratio": pick(headers, "Up Market Capture Ratio", "Up Market Capture\nRatio"),
                "Down Market Capture Ratio": pick(headers, "Down Market Capture Ratio", "Down Market Capture\nRatio"),
                "Maximum Drawdown": pick(headers, "Maximum Drawdown"),
                "R-Squared": pick(headers, "R-Squared", "R Squared"),
                "Information Ratio": pick(headers, "Information Ratio"),
            }
            missing = [k for k in required if h_idx.get(k, -1) < 0]
            if missing:
                print(f"Consolidated headers missing aliases for: {missing}")
            # iterate rows after detected header
            for r in range(header_row_idx + 1, sh.nrows):
                key = sh.cell_value(r, 0)
                if key in (None, ""):
                    continue
                key = str(key).strip()
                raw_row = [sh.cell_value(r, c) for c in range(sh.ncols)]
                metrics = build_stats_row(raw_row, h_idx)
                if key:
                    RISK_METRICS_BY_MFTOOLS_KEY[normalize_key(key)] = metrics
        else:
            wb = load_workbook(filename=path, data_only=True, read_only=True)
            ws = wb.active
            # buffer first 200 rows to locate header row
            buffered = []
            it = ws.iter_rows(values_only=True)
            for _ in range(200):
                try:
                    buffered.append(next(it))
                except StopIteration:
                    break
            header_row = None
            header_row_idx = None
            for i, row in enumerate(buffered):
                if not row:
                    continue
                joined = " ".join((str(v).strip().lower() if v is not None else "") for v in row)
                if "volatility" in joined and "sharpe" in joined:
                    header_row = row
                    header_row_idx = i
                    break
            if header_row is None:
                if buffered:
                    header_row = buffered[0]
                    header_row_idx = 0
                else:
                    header_row = []
                    header_row_idx = 0
            headers = [str(h).strip() if h is not None else "" for h in header_row]
            h_idx = {
                "Category": pick(headers, "Category"),
                "NAV": pick(headers, "NAV", "Nav"),
                "AUM": pick(headers, "AUM", "AUM (Crore)", "AUM (Cr)", "AUM (in Cr)"),
                "TER": pick(headers, "TER", "TER (%)"),
                "1 Yr Rtn": pick(headers, "1 Yr Rtn", "1 Year Rtn", "1 Yr Rtn (%)", "1 Year Rtn (%)"),
                "3 Yr Rtn": pick(headers, "3 Yr Rtn", "3 Year Rtn", "3 Yrs Rtn (%)", "3 Yr Rtn (%)"),
                "5 Yr Rtn": pick(headers, "5 Yr Rtn", "5 Year Rtn", "5 Yrs Rtn (%)", "5 Yr Rtn (%)"),
                "10 Yr Rtn": pick(headers, "10 Yr Rtn", "10 Year Rtn", "10 Yrs Rtn (%)", "10 Yr Rtn (%)"),
                "Volatility": pick(headers, "Volatility"),
                "Standard Deviation": pick(headers, "Standard Deviation"),
                "Sharpe Ratio": pick(headers, "Sharpe Ratio", "Sharp Ratio"),
                "Beta": pick(headers, "Beta"),
                "Alpha": pick(headers, "Alpha"),
                "Mean": pick(headers, "Mean"),
                "Sortino Ratio": pick(headers, "Sortino Ratio"),
                "Up Market Capture Ratio": pick(headers, "Up Market Capture Ratio", "Up Market Capture\nRatio"),
                "Down Market Capture Ratio": pick(headers, "Down Market Capture Ratio", "Down Market Capture\nRatio"),
                "Maximum Drawdown": pick(headers, "Maximum Drawdown"),
                "R-Squared": pick(headers, "R-Squared", "R Squared"),
                "Information Ratio": pick(headers, "Information Ratio"),
            }
            missing = [k for k in required if h_idx.get(k, -1) < 0]
            if missing:
                print(f"Consolidated headers missing aliases for: {missing}")
            # process buffered rows after header
            for r in buffered[header_row_idx + 1:]:
                if not r:
                    continue
                key = r[0]
                if key is None or str(key).strip() == "":
                    continue
                key = str(key).strip()
                metrics = build_stats_row(list(r), h_idx)
                if key:
                    RISK_METRICS_BY_MFTOOLS_KEY[normalize_key(key)] = metrics
            # continue with the rest of rows from iterator
            for r in it:
                if not r:
                    continue
                key = r[0]
                if key is None or str(key).strip() == "":
                    continue
                key = str(key).strip()
                metrics = build_stats_row(list(r), h_idx)
                if key:
                    RISK_METRICS_BY_MFTOOLS_KEY[normalize_key(key)] = metrics
    except Exception as ex:
        print(f"Failed to load consolidated stats from {path}: {ex}")

def enrich_from_mftools(valueDict, ak_name):
    try:
        mkey = MFT_AK_TO_KEY.get(ak_name)
        if isinstance(mkey, str) and mkey:
            metrics = RISK_METRICS_BY_MFTOOLS_KEY.get(normalize_key(mkey))
            if metrics:
                for k, v in metrics.items():
                    if v is not None:
                        valueDict[k] = v
    except Exception as ex:
        print(f"Consolidated stats enrichment failed for {ak_name}: {ex}")

def get_stock_prices(tickers):
    trendDetails = list()
    futures = list()
    symbols_no_data = []
    if not tickers:
        return trendDetails

    # Use ThreadPoolExecutor for I/O-bound web requests
    with concurrent.futures.ThreadPoolExecutor(max_workers=min(16, len(tickers))) as executor:
        for count, ticker in enumerate(tickers):
            print(count + 1, ticker)
            futures.append(executor.submit(get_stock_info, ticker))

        for future in concurrent.futures.as_completed(futures):
            try:
                has_data, value = future.result()
            except Exception as ex:
                has_data, value = False, f"unhandled_error: {ex}"
            if has_data:
                trendDetails.append(value)
            else:
                symbols_no_data.append(value)

    # Update the global funds_with_no_data
    global funds_with_no_data
    funds_with_no_data = symbols_no_data

    return trendDetails

def get_stock_info(symbol):
    # symbol can be of the form "DisplayName:slug". Extract both parts.
    sym0 = symbol
    if ':' in symbol:
        sym0, _ = symbol.split(':', 1)
    valueDict = {'Fund': sym0}
    enrich_from_mftools(valueDict, sym0)
    has_consolidated = len(valueDict.keys()) > 1
    if not has_consolidated:
        return (False, sym0)

    enrich_category_and_benchmark_returns(valueDict, sym0)

    # Resolve scheme code via amfiKey mapping from mftools JSON (no web search)
    scheme_code = MFT_AK_TO_AMFI.get(sym0)
    if isinstance(scheme_code, str) and scheme_code.strip():
        valueDict['Scheme Code'] = scheme_code.strip()

    scheme_code = valueDict.get('Scheme Code')
    if isinstance(scheme_code, str) and scheme_code.strip():
        try:
            groww_page_url = f"https://groww.in/v1/api/data/mf/web/v1/scheme/portfolio/{scheme_code}/stats"
            gp_resp = requests.get(groww_page_url, timeout=20)
            if gp_resp.status_code == 200:
                data = json.loads(gp_resp.text)
                valueDict['P/E Ratio'] = data.get("pe")
                valueDict['P/B Ratio'] = data.get("pb")
        except Exception:
            # Do not fail overall parsing if Groww page structure changes
            pass

    return (True, valueDict)

def export_to_file(data):
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    csv_file_path = f"fund-stats_{timestamp}.csv"

    column_order = ['Fund',
                   'Category',
                   'Scheme Code',
                   'Total Assets (in Cr)',
                   'TER',
                   '1 Year CAGR',
                   '1 Year Category CAGR',
                   '1 Year Benchmark CAGR',
                   '3 Years CAGR',
                   '3 Years Category CAGR',
                   '3 Years Benchmark CAGR',
                   '5 Years CAGR',
                   '5 Years Category CAGR',
                   '5 Years Benchmark CAGR',
                   '10 Years CAGR',
                   '10 Years Category CAGR',
                   '10 Years Benchmark CAGR',
                   'Benchmark Type',
                   'NAV',
                   'Alpha',
                   'Beta',
                   'Standard Deviation',
                   'Sharpe Ratio',
                   "Volatility",
                   "Mean",
                   "Sortino Ratio", 
                   "Up Market Capture\nRatio", 
                   "Down Market Capture\nRatio", 
                   "Maximum Drawdown", 
                   "R-Squared",
                   "Information Ratio",
                   'P/E Ratio',
                   'P/B Ratio'
                   ]

    fundsByType = {}
    for fund_data in data:
        if 'Category' in fund_data:
            category = fund_data.get('Category')
            fundStats = []
            if category in fundsByType:
                fundStats = fundsByType.get(category)
            fundStats.append([fund_data.get(column, '') for column in column_order])
            fundsByType[category] = fundStats
        else:
            print("No category for fund " + str(fund_data) + " hence skipping.")

    with open(csv_file_path, 'w', newline='') as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(column_order)

        # Prefer configured category ordering when labels overlap, then append any remaining categories.
        preferred = [c for c in (MFT_CATEGORIES or []) if c in fundsByType]
        remaining = [c for c in fundsByType.keys() if c not in preferred]
        categories_to_write = preferred + remaining

        for category in categories_to_write:
            writer.writerow([category])
            writer.writerows(fundsByType[category])

def run(consolidated: str, benchmark_returns: str, category_monitor: str, mftools_json: str) -> None:
    # load mappings
    load_mftools_mapping(mftools_json)
    load_risk_ratios(consolidated)
    load_benchmark_returns(benchmark_returns)
    load_category_returns(category_monitor)

    # extract fund names from JSON (akKey list)
    funds = MFT_AK_LIST or []

    # extract data for funds
    extracted_data = get_stock_prices(funds)

    print(f"\033[91m{len(funds_with_no_data)} funds have no data. These are, {funds_with_no_data}.\033[0m")

    # export CSV file (existing behavior)
    export_to_file(extracted_data)


def main() -> None:
    parser = argparse.ArgumentParser(description="Advisor parser using consolidated stats + Groww P/E/P/B")
    project_root = os.getcwd()
    default_mftools_json = os.path.join(project_root, "resources", "funds_and_categories_with_mftools.json")
    default_consolidated = os.path.join(project_root, "consolidated-mft-returns.xls")
    default_benchmark_returns = os.path.join(project_root, "benchmark-returns.xls")
    default_category_monitor = os.path.join(project_root, "category-monitor.xls")
    parser.add_argument("--consolidated", default=default_consolidated, help="Path to consolidated-mft-returns Excel (.xls or .xlsx)")
    parser.add_argument("--risk-ratios", dest="consolidated", help=argparse.SUPPRESS)
    parser.add_argument("--benchmark-returns", default=default_benchmark_returns, help="Path to benchmark-returns.xls/.xlsx")
    parser.add_argument("--category-monitor", default=default_category_monitor, help="Path to category-monitor.xls/.xlsx")
    parser.add_argument("--mftools-json", default=default_mftools_json, help="Path to funds_and_categories_with_mftools.json (defaults to file alongside this script)")
    args = parser.parse_args()
    run(args.consolidated, args.benchmark_returns, args.category_monitor, args.mftools_json)


if __name__ == "__main__":
    main()
