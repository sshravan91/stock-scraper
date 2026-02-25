import argparse
import concurrent.futures
import json
import os
import re
from typing import Any

import requests
import xlrd
import xlwt
from openpyxl import load_workbook


def normalize(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


def load_rows(path: str) -> tuple[list[str], list[list[Any]]]:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        wb = xlrd.open_workbook(path)
        sh = wb.sheet_by_index(0)
        rows = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    else:
        wb = load_workbook(filename=path, data_only=True, read_only=True)
        ws = wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]

    if not rows:
        return [], []
    headers = [str(v or "").strip() for v in rows[0]]
    return headers, rows[1:]


def write_xls(path: str, headers: list[str], rows: list[list[Any]]) -> str:
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Augmented Returns")

    for c, h in enumerate(headers):
        ws.write(0, c, h)
    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row):
            ws.write(r, c, val if val is not None else "")

    wb.save(path)
    return os.path.abspath(path)


def build_alias_index(mapping_json_path: str) -> dict[str, dict[str, Any]]:
    with open(mapping_json_path, "r") as f:
        payload = json.load(f)

    index: dict[str, dict[str, Any]] = {}
    for rec in payload.get("funds", []):
        aliases = set()
        ak = rec.get("akKey")
        mk = rec.get("mftools_key")
        if isinstance(ak, str) and ak.strip():
            aliases.add(normalize(ak))
            aliases.add(normalize(ak.replace("-", " ")))
        if isinstance(mk, str) and mk.strip():
            aliases.add(normalize(mk))
        for alias in aliases:
            if alias:
                index[alias] = rec
    return index


def extract_scheme_code(scheme_name: str, alias_index: dict[str, dict[str, Any]]) -> str | None:
    rec = alias_index.get(normalize(scheme_name))
    if not rec:
        return None
    amfi = rec.get("amfiKey")
    if amfi is None:
        return None
    code = str(amfi).strip()
    if not code or code.lower() == "nan":
        return None
    return code


def fetch_groww_stats(scheme_code: str) -> tuple[Any, Any]:
    url = f"https://groww.in/v1/api/data/mf/web/v1/scheme/portfolio/{scheme_code}/stats"
    try:
        resp = requests.get(url, timeout=20)
        if resp.status_code != 200:
            return "", ""
        data = resp.json()
        return data.get("pe", ""), data.get("pb", "")
    except Exception:
        return "", ""


def augment_with_groww(input_path: str, mapping_json_path: str, output_path: str) -> str:
    headers, rows = load_rows(input_path)
    if not headers:
        raise RuntimeError(f"No rows found in input file: {input_path}")

    scheme_col = -1
    for i, h in enumerate(headers):
        if normalize(h) in ("schemename", "scheme", "fundname", "fund"):
            scheme_col = i
            break
    if scheme_col < 0:
        raise RuntimeError("Could not find 'Scheme name' column in input file")

    alias_index = build_alias_index(mapping_json_path)

    # Resolve scheme code for each row.
    scheme_codes: list[str | None] = []
    for row in rows:
        scheme_name = str(row[scheme_col] if scheme_col < len(row) else "").strip()
        scheme_codes.append(extract_scheme_code(scheme_name, alias_index))

    # Fetch Groww metrics concurrently.
    results: list[tuple[Any, Any]] = [("", "") for _ in rows]
    max_workers = min(16, max(1, len(rows)))
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {}
        for idx, code in enumerate(scheme_codes):
            if code:
                future_map[executor.submit(fetch_groww_stats, code)] = idx
        for future in concurrent.futures.as_completed(future_map):
            idx = future_map[future]
            try:
                results[idx] = future.result()
            except Exception:
                results[idx] = ("", "")

    # Append columns.
    out_headers = headers + ["P/E Ratio", "P/B Ratio"]
    out_rows: list[list[Any]] = []
    for row, (pe, pb) in zip(rows, results):
        out_rows.append(list(row) + [pe, pb])

    return write_xls(output_path, out_headers, out_rows)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Augment consolidated MFT returns with Groww P/E and P/B ratios"
    )
    script_dir = os.path.dirname(os.path.abspath(__file__))
    parser.add_argument(
        "--input",
        default=os.path.join(script_dir, "consolidated-mft-returns.xls"),
        help="Path to consolidated-mft-returns.xls/.xlsx",
    )
    parser.add_argument(
        "--mapping-json",
        default=os.path.join(script_dir, "funds_and_categories_with_mftools.json"),
        help="Path to funds_and_categories_with_mftools.json",
    )
    parser.add_argument(
        "--out",
        default=os.path.join(script_dir, "consolidated-mft-returns-augmented.xls"),
        help="Output XLS path",
    )
    args = parser.parse_args()

    saved = augment_with_groww(args.input, args.mapping_json, args.out)
    print(f"✅ Augmented file saved to: {saved}")


if __name__ == "__main__":
    main()
